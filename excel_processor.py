"""
This is the main excel processor file.
"""
# ## features:

#     1- process excel files. (input/output)

#     2- reporting category per gender.

#     3- report monthly profit

#     4- report quarterly profit

#     5- most profitable categories

#     6*- RFM for customer conversion

#     6- most profitable customer types

#     7- most bought categories normal customer
import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.chart import BarChart, Reference


def read_process_xl(filename: str) -> pd.DataFrame:
    df: pd.DataFrame = pd.read_excel(filename)
    processed_df = df[["Gender", "Product line", "Total"]]
    return processed_df


def write_xl(data: pd.DataFrame) -> None:
    data.to_excel(r"output\\report_category_per_gender_20220406.xlsx",
                  sheet_name="Report", startrow=4)


def analyze_category_per_gender(data: pd.DataFrame) -> pd.DataFrame:
    report: pd.DataFrame = df.pivot_table(
        index="Gender", columns="Product line", values="Total", aggfunc="sum").round(0)
    return report


def monthly_profit(data):
    pass


def quarter_profit(data):
    pass


def most_profitable_category(data):
    pass


def most_profitable_customer_type(data):
    pass


def most_bought_categories_normal_customer(data):
    pass


df: pd.DataFrame = read_process_xl(r'resources\\supermarket_sales.xlsx')
report: pd.DataFrame = analyze_category_per_gender(df)
write_xl(report)

workbook: Workbook = load_workbook(r"output\\report_category_per_gender_20220406.xlsx")
sheet = workbook["Report"]

min_col = workbook.active.min_column
max_col = workbook.active.max_column
min_row = workbook.active.min_row
max_row = workbook.active.max_row

data = Reference(sheet, min_col=min_col+1, max_col=max_col,
                 min_row=min_row, max_row=max_row)
categories = Reference(sheet, min_col=min_col,
                       max_col=max_col, min_row=min_row+1, max_row=max_row)

bar_chart = BarChart()

bar_chart.add_data(data, titles_from_data=True)
bar_chart.set_categories(categories)

sheet.add_chart(bar_chart, "B20")

workbook.save("report_category_per_gender_20220406.xlsx")
